#!/usr/bin/env python3
"""人時分析 xlsx → 鮮湧的人時分析頁面靜態資料。

⚠️ 產出要貼進**兩個檔**（只更新一邊，同一份報告在手機與桌面會顯示不同數字）：
   · web/public/shianyong-labour.html   ← 手機版 · 圖文選單「報工單」的目的地
   · web/public/shianyong-warroom.html  ← 桌面戰情室的「人時分析」分頁


用法：
    python3 scripts/labour-hours-probe.py ~/Downloads/人時分析每日報告_模擬30日_20260716-20260826.xlsx
    python3 scripts/labour-hours-probe.py <xlsx> --emit-js

⚠️⚠️ **這份是模擬資料。** 檔案自己在彙總表最後一列寫著：
   「本檔為模擬資料，僅日期 20260826 對應原始報表；其餘 29 日依工位特性、
     星期效應與設定之特殊事件日隨機產生」
   → 30 天裡**只有 1 天是真的**。畫面上必須講出來（同巡檢那頁的「快照」橫幅）。

⚠️ 兩個讀檔陷阱：
   1. **資料從 B 欄開始**，A 欄整欄空白 —— `values_only=True` 的 tuple 索引要 +1，
      照 A 欄起算會全部讀成 None（第一版就是這樣，讀出 0 列）。
   2. 時間欄是 `datetime.timedelta` 不是 `time` —— 超過 24 小時的「四工位合計」
      會變成 `1 day, 0:35:14`，用 `.hour` 取值會 AttributeError。
"""
import datetime
import json
import os
import statistics
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("需要 openpyxl：pip install openpyxl", file=sys.stderr)
    raise SystemExit(1)

SLOT_LABEL = {"上午": "am", "下午": "pm", "晚上": "night"}


def secs(v) -> int | None:
    """時間欄 → 秒。

    ⚠️ **同一個檔案裡有三種型別**：
      · 彙總表的累積時間 → `timedelta`（超過 24h 會是 `1 day, 0:35:14`）
      · 單日表的時段小計 → `timedelta`
      · 單日表的**逐事件時間 → 純字串 `"00:45:53"`**
    第一版只認前兩種，事件明細整批讀成 None → 事件總數 0。
    而且**它不會報錯**，只是靜靜地少掉 900 筆。
    """
    if isinstance(v, datetime.timedelta):
        return int(v.total_seconds())
    if isinstance(v, datetime.time):
        return v.hour * 3600 + v.minute * 60 + v.second
    if isinstance(v, str) and v.count(":") == 2:
        try:
            h, m, sec = (int(x) for x in v.split(":"))
            return h * 3600 + m * 60 + sec
        except ValueError:
            return None
    return None


def read_summary(wb) -> tuple[dict, list[dict]]:
    ws = wb["彙總"]
    rows = list(ws.iter_rows(values_only=True))

    cap = None
    disclaimer = None
    period = None
    for r in rows:
        for c in r:
            if not isinstance(c, str):
                continue
            if c.startswith("期間："):
                period = c.replace("期間：", "").strip()
            if "模擬資料，僅日期" in c:
                disclaimer = c
        if r[1] == "每日可用時段上限":
            cap = secs(r[2])

    daily = []
    for r in rows:
        if not isinstance(r[1], datetime.datetime):
            continue
        st = [secs(r[i]) for i in range(3, 7)]
        if any(v is None for v in st):
            continue
        daily.append({
            "d": r[1].strftime("%Y-%m-%d"),
            "dow": r[2],
            "s": st,
            "total": secs(r[7]),
            "rate": round(float(r[8]), 4) if r[8] is not None else None,
            "note": r[9] or "",
        })
    meta = {"period": period, "capPerStation": cap, "disclaimer": disclaimer}
    return meta, daily


def read_day(wb, sheet: str) -> dict:
    """單日表：三個時段的累積 + 逐事件明細。"""
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    slots: dict[str, list[int | None]] = {}
    events: list[dict] = []
    for r in rows:
        label = r[1] if isinstance(r[1], str) else ""
        for zh, key in SLOT_LABEL.items():
            if label.startswith(f"{zh}累積時間"):
                slots[key] = [secs(r[i]) for i in range(2, 6)]
        # 事件明細列：事件編號是 8 位數字
        # 事件編號是 10 位數字（YYYYMMDDnn）· 是**字串**不是 int
        if isinstance(r[1], (int, str)) and str(r[1]).isdigit() and len(str(r[1])) == 10:
            d = secs(r[3])
            if d is not None:
                events.append({"id": str(r[1]), "st": r[2], "sec": d})
    return {"slots": slots, "events": events}


def main() -> int:
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    path = args[0] if args else os.path.expanduser(
        "~/Downloads/人時分析每日報告_模擬30日_20260716-20260826.xlsx")
    wb = load_workbook(path, data_only=True)
    meta, daily = read_summary(wb)
    if not daily:
        print("❌ 彙總表讀不到任何一天（檢查是不是又從 A 欄起算了）", file=sys.stderr)
        return 1

    days = {d["d"]: read_day(wb, d["d"].replace("-", "")) for d in daily
            if d["d"].replace("-", "") in wb.sheetnames}
    for d in daily:
        k = d["d"]
        d["slots"] = days.get(k, {}).get("slots", {})
        d["nEvents"] = len(days.get(k, {}).get("events", []))

    # 只帶「真實那一天」的逐事件明細 —— 其餘 29 天是隨機產生的，
    # 把假事件一筆筆列出來只會讓人以為那是實測（R11）。
    real_key = None
    if meta["disclaimer"]:
        import re
        m = re.search(r"僅日期 (\d{8}) 對應原始報表", meta["disclaimer"])
        if m:
            real_key = f"{m.group(1)[:4]}-{m.group(1)[4:6]}-{m.group(1)[6:]}"
    real_events = days.get(real_key, {}).get("events", []) if real_key else []

    if "--emit-js" in sys.argv:
        st_avg = [round(statistics.mean(d["s"][i] for d in daily)) for i in range(4)]
        slot_avg = {k: [round(statistics.mean(d["slots"][k][i] for d in daily if d["slots"].get(k)))
                        for i in range(4)] for k in ("am", "pm", "night")}
        out = {
            "meta": {
                **meta,
                "days": len(daily),
                "realDay": real_key,
                "stations": ["S01", "S02", "S03", "S04"],
                "avgRate": round(statistics.mean(d["rate"] for d in daily), 4),
                "stationAvg": st_avg,
                "slotAvg": slot_avg,
                "totalEvents": sum(d["nEvents"] for d in daily),
                "generatedFrom": "scripts/labour-hours-probe.py --emit-js",
            },
            "daily": daily,
            "realEvents": real_events,
        }
        print("/* ⚠️ 自動產生 —— 不要手改。\n"
              "   來源：scripts/labour-hours-probe.py --emit-js <xlsx>\n"
              "   ⚠️ **模擬資料**：30 天裡只有 " + (real_key or "?") + " 對應原始報表。*/")
        print("const LABOUR=" + json.dumps(out, ensure_ascii=False, separators=(",", ":")) + ";")
        return 0

    print(f"═══ 人時分析 · {path}")
    print(f"    {meta['period']} · {len(daily)} 個工作日 · 每工位每日上限 {meta['capPerStation']/3600:.2f}h")
    print(f"\n⚠️ {meta['disclaimer']}\n")
    print(f"{'日期':<12}{'週':<4}" + "".join(f"{s:>9}" for s in ("S01", "S02", "S03", "S04"))
          + f"{'稼動率':>9}  備註")
    for d in daily:
        print(f"{d['d']:<12}{d['dow']:<4}"
              + "".join(f"{v/3600:>8.2f}h" for v in d["s"])
              + f"{d['rate']:>8.1%}  {d['note']}")
    print(f"\n平均稼動率 {statistics.mean(d['rate'] for d in daily):.1%}")
    for i, s in enumerate(("S01", "S02", "S03", "S04")):
        v = [d["s"][i] for d in daily]
        print(f"  {s}  平均 {statistics.mean(v)/3600:5.2f}h   {min(v)/3600:.2f} ~ {max(v)/3600:.2f}h")
    print(f"\n事件總數 {sum(d['nEvents'] for d in daily)} · 真實那天（{real_key}）{len(real_events)} 筆")
    return 0


if __name__ == "__main__":
    sys.exit(main())
